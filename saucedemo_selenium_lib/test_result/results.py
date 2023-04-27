""" Test Results

Contains classes or functions for parsing the test results and generating test results table
"""
from collections import OrderedDict
from lxml.html import parse
import os

from openpyxl import Workbook
import openpyxl.styles

boldfont = openpyxl.styles.Font(bold=True)
titlefont = openpyxl.styles.Font(bold=True, size=16)
titlealignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")


class TargetTestResults:
    """Test results for a target"""

    def __init__(self, target_name, passes=None, failures=None, target_id=None):

        self._target_name= target_name
        if target_id:
            self._target_id = target_id
        else:
            self._target_id = target_name

        if passes is None:
            passes = 0
        if failures is None:
            failures = 0
        self._passes = passes
        self._failures = failures
        self._results = OrderedDict()
        self._percent = None
        self._total_tests_executed = None

    def __str__(self):
        return f"({self.target_name},Total Tests: {self.total_tests_executed},  passes: {self.passes}, failures: {self.failures}, pass rate: {self.percent})"


    @property
    def target_name(self):
        return self._target_name

    @property
    def results(self):
        return self._results

    @property
    def passes(self):
        return self._passes

    @property
    def failures(self):
        return self._failures

    @property
    def total_tests_executed(self):
        if self._total_tests_executed:
            return self._total_tests_executed
        self._total_tests_executed = self._passes + self._failures
        return self._total_tests_executed

    @property
    def percent(self):
        if self._percent:
            return self._percent
        else:
            total = self._passes + self._failures
            percent = 0
            if total > 0:
                percent = round(float((self._passes / total) * 100), 1)
            return percent

    def add_test_case_result(self, test_case, test: dict):
        """Add test result of a specific test case"""
        if test_case not in self._results:
            self._results[test_case] = []
        self._results[test_case].append(test)


class HTMLTestResultsParser:
    """Parse test results from html generated during tests by pytest-html"""

    def __init__(self, target_name, file_path):
        self._target_name = target_name
        self._file = file_path
        print(f"Parsing file: {self._file}")
        self._parser = parse(self._file).getroot()
        self._results = TargetTestResults(
            self._target_name, self.get_total_passed(), self.get_total_failed()
        )

    @property
    def html_report_file_path(self):
        return self._file

    def _get_elements_by_xpath(self, xpath: str):
        elements = self._parser.xpath(xpath)
        return elements

    def _get_text_content(self, element) -> str:
        text = element.text_content()
        return text

    def get_total_passed(self, xpath="//span[@class='passed']"):
        """Get total number of tests that passed"""
        element = self._get_elements_by_xpath(xpath)[0]
        content = self._get_text_content(element)
        total = int(content.split(" ")[0])
        print(f"Total number of passed tests: {total}")
        return total

    def get_total_failed(
        self,
        failed_xpath="//span[@class='failed']",
        errors_xpath="//span[@class='error']",
    ):
        """Get total number of tests that passed"""
        failures_element = self._get_elements_by_xpath(failed_xpath)[0]
        errors_element = self._get_elements_by_xpath(errors_xpath)[0]
        failures = self._get_text_content(failures_element)
        errors = self._get_text_content(errors_element)
        total = int(failures.split(" ")[0]) + int(errors.split(" ")[0])
        print(f"Total number of failed tests: {total}")
        return total

    def get_tests_results(
        self, xpath="//tbody[contains(@class,'results-table-row')]"
    ) -> TargetTestResults:
        elements = self._get_elements_by_xpath(xpath)

        for element in elements:
            detail = self._get_text_content(element.find_class("col-name")[0]).split(
                "::"
            )
            # print(f"test Name: {detail}")

            test_case = detail[1]

            test_name = detail[2].split("test_")[1].replace("_", " ")
            result = self._get_text_content(element.find_class("col-result")[0])
            test = {"name": test_name, "result": result}
            self._results.add_test_case_result(test_case=test_case, test=test)
        return self._results

class ResultsTableCreator:
    """Create results table excel for test results"""

    def __init__(self, test_results: [TargetTestResults], output_path=""):
        self._test_results = test_results
        self._output_path = output_path
        self._wb = Workbook()
        self._wb.remove(self._wb.active)

    def _create_sheet_with_title(self, title):
        sheet = self._wb.create_sheet(title)
        sheet.column_dimensions["A"].width = 100
        sheet.title = title
        sheet.merge_cells("A1:B1")
        sheet["A1"].value = title
        sheet["A1"].font = titlefont
        sheet["A1"].alignment = titlealignment
        return sheet

    # { <use case name>: [(<test 1 name>, <test 1 status>), ... ] ... }
    def create_results_table(self, file_name_phrase="all"):
        print("Target tests.complete. Writing results table ...")

        overview_sheet = self._create_sheet_with_title("Overview")
        overview_sheet["A2"].value = "Name"
        overview_sheet["B2"].value = "Succeeded"
        overview_sheet["C2"].value = "Failed"
        overview_sheet["D2"].value = "Success ratio"
        overview_sheet["A2"].font = boldfont
        overview_sheet["B2"].font = boldfont
        overview_sheet["C2"].font = boldfont
        overview_sheet["D2"].font = boldfont

        success_counter = 0
        failure_counter = 0

        overview_iter = 3

        for target in self._test_results:

            self._create_service_result_sheet(target)

            overview_sheet.cell(
                row=overview_iter, column=1
            ).value = target.target_name
            overview_sheet.cell(row=overview_iter, column=2).value = target.passes
            overview_sheet.cell(row=overview_iter, column=3).value = target.failures

            if target.passes + target.failures > 0:
                overview_sheet.cell(row=overview_iter, column=4).value = round(
                    float(target.passes) * 100 / (target.passes + target.failures), 1
                )
            overview_iter += 1
            success_counter += target.passes
            failure_counter += target.failures

        overview_sheet.cell(row=overview_iter + 1, column=1).value = "Overall"
        overview_sheet.cell(row=overview_iter + 1, column=2).value = success_counter
        overview_sheet.cell(row=overview_iter + 1, column=3).value = failure_counter
        total_counter = failure_counter + success_counter
        total_percent = 0
        if total_counter > 0:
            total_percent = round(float(success_counter / total_counter) * 100, 2)
        overview_sheet.cell(row=overview_iter + 1, column=4).value = total_percent

        overview_sheet.cell(row=overview_iter + 1, column=1).font = boldfont
        overview_sheet.cell(row=overview_iter + 1, column=2).font = boldfont
        overview_sheet.cell(row=overview_iter + 1, column=3).font = boldfont
        overview_sheet.cell(row=overview_iter + 1, column=4).font = boldfont
        file_path = os.path.join(
            self._output_path, f"table-test-results-{file_name_phrase}.xlsx"
        )
        self._wb.save(file_path)

        print("The workbook was successfully created.")
        print(f"Excel File Path:- {file_path}")

    def _create_service_result_sheet(self, target: TargetTestResults):
        """Create result sheet for a service where its detail test result content is written"""
        sheet = self._create_sheet_with_title(target.target_name)

        sheet["A3"].value = "Summary"
        sheet["A3"].font = boldfont

        sheet["A4"].value = "Total Test Executed:"
        sheet["B4"].value = target.total_tests_executed
        sheet["B4"].font = boldfont

        sheet["A5"].value = "Total Passes:"
        sheet["B5"].value = target.passes
        sheet["B5"].font = boldfont

        sheet["A6"].value = "Total Failures:"
        sheet["B6"].value = target.failures
        sheet["B6"].font = boldfont

        sheet["A7"].value = "Pass Ratio:"
        sheet["B7"].value = target.percent
        sheet["B7"].font = boldfont

        sheet["A9"].value = "Test Result Details"
        sheet["A9"].font = boldfont
        line_iter = 10

        for name, tests in target.results.items():
            print(f"Use case: {name}")
            print(f"Tests: {tests}")
            sheet.cell(row=line_iter, column=1).value = name
            sheet.cell(row=line_iter, column=1).font = boldfont
            line_iter += 1
            for test in tests:
                if test["result"] != "Skipped":
                    result = test["result"]
                    if result == "Error":
                        result = "Failed"
                    sheet.cell(row=line_iter, column=1).value = test["name"]
                    sheet.cell(row=line_iter, column=2).value = result
                    line_iter += 1